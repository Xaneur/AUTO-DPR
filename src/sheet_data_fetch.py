import openpyxl
from openpyxl.utils import get_column_letter
import datetime
from typing import List, Optional
from utils.logger import get_logger
logger = get_logger(__name__)

def get_available_sheets(file_path: str) -> List[str]:
    """
    Get a list of all sheet names from the Excel file.
    
    Args:
        file_path (str): Path to the Excel file
        
    Returns:
        List[str]: List of sheet names
        
    Raises:
        FileNotFoundError: If the specified file doesn't exist
        Exception: For other errors during file processing
    """
    try:
        logger.info(f"Fetching available sheets from: {file_path}")
        # Load the workbook in read-only mode for better performance
        wb = openpyxl.load_workbook(file_path, read_only=True)
        sheets = [sheet.strip() for sheet in wb.sheetnames if sheet.strip()]  # Remove empty or whitespace-only names
        wb.close()  # Important to close the workbook when done
        logger.info(f"Found {len(sheets)} sheets: {', '.join(sheets)}")
        return sheets
    except FileNotFoundError as e:
        logger.error(f"File not found: {file_path}")
        raise FileNotFoundError(f"The specified file was not found: {file_path}") from e
    except Exception as e:
        logger.error(f"Error reading sheets from {file_path}: {str(e)}")
        raise Exception(f"Failed to read sheets from {file_path}: {str(e)}") from e

def get_descriptions_with_index(file_path, sheet_name="July.25"):
    """
    Extract Description column (C) data with row indices.
    Returns string format: "[(row_index, description), (row_index, description), ...]"
    """
    # Load the workbook
    logger.info(f"getting descriptions from workbook from path : {file_path} and sheet name is : {sheet_name}")
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name]
    
    # Create list of tuples (row_index, description)
    descriptions = []
    
    # Start from row 2 (skip header) and go to the last row
    for row_num in range(5, ws.max_row + 1):
        cell_value = ws[f"C{row_num}"].value
        
        # Only include non-empty values
        if cell_value is not None and str(cell_value).strip() != "":
            descriptions.append((row_num, cell_value))
    
    return str(descriptions)

def get_date_column(file_path, sheet_name="July.25", date: datetime.date = None):
    """
    find the column containing today's date in the first row.
    """
    logger.info(f"getting date column from workbook from path : {file_path} and sheet name is : {sheet_name}")
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name]
    
    for cell in ws[1]:
        if isinstance(cell.value, datetime.datetime):
            if date is None:
                if cell.value.date() == datetime.date.today():
                    logger.info(f"date column is : {cell.column} and col name is : {get_column_letter(cell.column)}")
                    return cell.column+1
            else:
                if cell.value.date() == date:
                    logger.info(f"date column is : {cell.column} and col name is : {get_column_letter(cell.column)}")
                    return cell.column+1
    
    logger.info("date column not found")
    return None

def put_logs_in_file(file_path: str, sheet_name="LOGS", description=None, 
                   row_index=None, column_index=None, value: float = None,
                   name: str = None, location: str = None):
    """
    Log an entry to the specified sheet with additional metadata.
    
    Args:
        file_path (str): Path to the Excel file
        sheet_name (str): Name of the sheet to log to (default: "LOGS")
        description (str): Description of the log entry
        row_index (int): Row number that was updated
        column_index (int): Column number that was updated
        value: The value that was written
        name (str): Name of the person making the update
        location (str): Location where the update was made
    """
    wb = openpyxl.load_workbook(file_path)

    # Create the LOG sheet on first run and add headers
    if "LOGS" not in wb.sheetnames:
        ws = wb.create_sheet("LOGS")
        headers = [
            'Logged_At', 'Updated_Sheet', 'Name', 'Location', 
            'Description', 'Row', 'Column', 'Value'
        ]
        ws.append(headers)
    else:
        ws = wb["LOGS"]

    # Find first empty row (after any header)
    next_row = ws.max_row + 1
    # A safer check in case of trailing blanks:
    while ws.cell(row=next_row, column=1).value not in (None, ""):
        next_row += 1

    # Write the log entry
    ws.cell(row=next_row, column=1, value=datetime.datetime.now())
    ws.cell(row=next_row, column=2, value=sheet_name)
    ws.cell(row=next_row, column=3, value=name)
    ws.cell(row=next_row, column=4, value=location)
    ws.cell(row=next_row, column=5, value=description)
    ws.cell(row=next_row, column=6, value=row_index)
    ws.cell(row=next_row, column=7, value=column_index)
    ws.cell(row=next_row, column=8, value=value)

    wb.save(file_path)
    logger.info(f"log row {next_row} written successfully")

def update_sheet(file_path: str = "/Users/devrajsinhgohil/Desktop/DPR/excel_files/DPR.xlsx", 
               sheet_name: str = "July.25", 
               row_index: int = None, 
               column_index: int = None, 
               value: float = None) -> None:
    """
    Update a cell in the Excel sheet, adding the new value to any existing value.
    
    Args:
        file_path: Path to the Excel file
        sheet_name: Name of the sheet to update
        row_index: 1-based row index
        column_index: 1-based column index
        value: The value to add to the existing cell value
        
    Raises:
        ValueError: If required parameters are missing or invalid
    """
    logger.info(f"Updating sheet: {file_path}, sheet: {sheet_name}")
    logger.info(f"Row: {row_index}, Column: {column_index}, Adding value: {value}")
    
    if row_index is None or column_index is None or value is None:
        raise ValueError("row_index, column_index, and value must be provided")
    
    if not isinstance(value, (int, float)):
        raise ValueError("Value must be a number")
    
    try:
        # Load the workbook
        wb = openpyxl.load_workbook(file_path)
        ws = wb[sheet_name]
        
        # Get the target cell
        cell = ws.cell(row=row_index, column=column_index)
        
        # Get the current value, defaulting to 0 if empty or not a number
        current_value = 0
        if cell.value is not None:
            try:
                current_value = float(cell.value)
            except (ValueError, TypeError):
                logger.warning(f"Existing value '{cell.value}' in cell {row_index},{column_index} is not a number. Treating as 0.")
        
        # Calculate new value by adding to existing
        new_value = current_value + value
        
        # Update the cell
        cell.value = new_value
        
        # Save the workbook
        wb.save(file_path)
        logger.info(f"Successfully updated cell {row_index},{column_index} with value: {new_value} (previous: {current_value}, added: {value})")
        
    except Exception as e:
        logger.error(f"Error updating sheet: {str(e)}")
        raise

if __name__ == "__main__":
    file_path = "/Users/devrajsinhgohil/Desktop/DPR/excel_files/DPR.xlsx"

    put_logs_in_file(file_path=file_path, sheet_name="LOGS", description="test", row_index=46, column_index=5, value=4.5)
    # sheet_name = "July.25"
    # result = get_descriptions_with_index(file_path)
    # print(result)

    # print("_" * 100)
    
    # date_column = get_date_column(file_path)
    # print(date_column)

    # update_sheet(file_path=file_path, sheet_name=sheet_name, row_index=46, column_index=5, value=4.5)
